from django.http import JsonResponse
from django.views.decorators.http import require_http_methods
from django.contrib.auth.decorators import login_required
import json
import openpyxl
from io import BytesIO

from accounts.models import Region, District, Mahalla
from tasks.models import Task, TaskAssignment, TaskResponse, TaskColumn, TaskQuestion


@login_required
def get_districts(request):
    """Viloyat bo'yicha tumanlar"""
    region_id = request.GET.get('region')

    if region_id:
        districts = District.objects.filter(
            region_id=region_id,
            is_active=True
        ).values('id', 'name')
        return JsonResponse(list(districts), safe=False)

    return JsonResponse([], safe=False)


@login_required
def get_mahallas(request):
    """Tuman bo'yicha mahallalar"""
    district_id = request.GET.get('district')

    if district_id:
        mahallas = Mahalla.objects.filter(
            district_id=district_id,
            is_active=True
        ).values('id', 'name')
        return JsonResponse(list(mahallas), safe=False)

    return JsonResponse([], safe=False)


@login_required
@require_http_methods(["POST"])
def task_auto_save(request, pk):
    """
    Vazifa javoblarini avtomatik saqlash
    Ko'p qatorli jadval to'liq qo'llab-quvvatlanadi
    """
    try:
        assignment = TaskAssignment.objects.select_related('task').get(
            pk=pk,
            leader=request.user
        )

        if not assignment.can_edit:
            return JsonResponse({
                'success': False,
                'error': 'Tahrirlash mumkin emas'
            })

        task = assignment.task
        saved_count = 0
        errors = []

        # ══════════════════════════════════════════════════════════════════════
        # TABLE TYPE - Ko'p qatorli jadval
        # ══════════════════════════════════════════════════════════════════════
        if task.type == Task.Type.TABLE:
            columns = {str(col.pk): col for col in task.columns.all()}

            for key, value in request.POST.items():
                if not key.startswith('col_'):
                    continue

                # col_UUID_ROWINDEX formatini parse qilish
                # Masalan: col_560c853b-4700-4224-9eb5-6f4a69a03a13_2
                parts = key.split('_')

                if len(parts) < 3:
                    continue

                try:
                    # UUID o'rtada, row_index oxirida
                    # parts = ['col', '560c853b-4700-4224-9eb5-6f4a69a03a13', '2']
                    column_id = '_'.join(parts[1:-1])  # UUID ni to'liq olish
                    row_index = int(parts[-1])  # Oxirgi element - row index

                    column = columns.get(column_id)
                    if not column:
                        continue

                    # Bo'sh qiymatlarni ham saqlash (o'chirish uchun)
                    if value == '' or value is None:
                        # Bo'sh bo'lsa, mavjud response ni o'chiramiz
                        TaskResponse.objects.filter(
                            assignment=assignment,
                            column=column,
                            row_index=row_index
                        ).delete()
                    else:
                        # Yangi qiymat saqlash
                        response, created = TaskResponse.objects.update_or_create(
                            assignment=assignment,
                            column=column,
                            row_index=row_index,
                            defaults={'question': None}
                        )
                        response.set_value(value)
                        saved_count += 1

                except (ValueError, IndexError) as e:
                    errors.append(f"Xatolik: {key} - {str(e)}")
                    continue

        # ══════════════════════════════════════════════════════════════════════
        # SURVEY TYPE - So'rovnoma
        # ══════════════════════════════════════════════════════════════════════
        elif task.type == Task.Type.SURVEY:
            questions = {str(q.pk): q for q in task.questions.all()}

            for key, value in request.POST.items():
                if not key.startswith('q_'):
                    continue

                try:
                    question_id = key[2:]  # 'q_' ni olib tashlash
                    question = questions.get(question_id)

                    if not question:
                        continue

                    # Multiple choice (checkbox) uchun
                    if question.answer_type == 'multiple':
                        values = request.POST.getlist(key)
                        response, created = TaskResponse.objects.update_or_create(
                            assignment=assignment,
                            question=question,
                            row_index=0,
                            defaults={'column': None}
                        )
                        response.value_json = values
                        response.save()
                    else:
                        if value == '' or value is None:
                            TaskResponse.objects.filter(
                                assignment=assignment,
                                question=question,
                                row_index=0
                            ).delete()
                        else:
                            response, created = TaskResponse.objects.update_or_create(
                                assignment=assignment,
                                question=question,
                                row_index=0,
                                defaults={'column': None}
                            )
                            response.set_value(value)
                    saved_count += 1

                except Exception as e:
                    errors.append(f"Savol xatoligi: {key} - {str(e)}")
                    continue

        # ══════════════════════════════════════════════════════════════════════
        # REPORT TYPE - Hisobot
        # ══════════════════════════════════════════════════════════════════════
        elif task.type == Task.Type.REPORT:
            report_text = request.POST.get('report_text', '')

            if report_text:
                response, created = TaskResponse.objects.update_or_create(
                    assignment=assignment,
                    column=None,
                    question=None,
                    row_index=0
                )
                response.value_text = report_text
                response.save()
                saved_count += 1

        # Progress yangilash
        assignment.update_progress()

        return JsonResponse({
            'success': True,
            'saved_count': saved_count,
            'progress': assignment.progress,
            'errors': errors if errors else None
        })

    except TaskAssignment.DoesNotExist:
        return JsonResponse({
            'success': False,
            'error': 'Vazifa topilmadi'
        }, status=404)
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)


@login_required
@require_http_methods(["POST"])
def task_excel_import(request, pk):
    """
    Excel fayldan ma'lumotlarni import qilish
    Yetakchi tayyor Excel yuklasa, tizim avtomatik jadvalga joylaydi
    """
    try:
        assignment = TaskAssignment.objects.select_related('task').get(
            pk=pk,
            leader=request.user
        )

        if not assignment.can_edit:
            return JsonResponse({
                'success': False,
                'error': 'Tahrirlash mumkin emas'
            })

        task = assignment.task

        if task.type != Task.Type.TABLE:
            return JsonResponse({
                'success': False,
                'error': 'Faqat jadval turidagi vazifalar uchun'
            })

        # Fayl olish
        excel_file = request.FILES.get('file')
        if not excel_file:
            return JsonResponse({
                'success': False,
                'error': 'Fayl yuklanmadi'
            })

        # Fayl turini tekshirish
        if not excel_file.name.endswith(('.xlsx', '.xls')):
            return JsonResponse({
                'success': False,
                'error': 'Faqat Excel fayl (.xlsx, .xls) yuklang'
            })

        # Excel ni o'qish
        try:
            wb = openpyxl.load_workbook(BytesIO(excel_file.read()))
            ws = wb.active
        except Exception as e:
            return JsonResponse({
                'success': False,
                'error': f'Excel faylni o\'qishda xatolik: {str(e)}'
            })

        # Ustunlarni olish
        columns = list(task.columns.order_by('order'))
        if not columns:
            return JsonResponse({
                'success': False,
                'error': 'Vazifada ustunlar yo\'q'
            })

        # Avvalgi javoblarni o'chirish
        assignment.responses.filter(column__isnull=False).delete()

        # Excel dan ma'lumotlarni o'qish
        imported_rows = 0
        errors = []

        # Birinchi qator - sarlavha, 2-qatordan boshlab ma'lumotlar
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True)):
            # Bo'sh qatorni o'tkazib yuborish
            if all(cell is None or cell == '' for cell in row):
                continue

            # Maksimal qatorlar tekshiruvi
            if imported_rows >= task.max_rows:
                errors.append(f"Maksimal {task.max_rows} ta qator ruxsat etilgan")
                break

            for col_idx, column in enumerate(columns):
                if col_idx < len(row):
                    value = row[col_idx]

                    if value is not None and value != '':
                        try:
                            response = TaskResponse.objects.create(
                                assignment=assignment,
                                column=column,
                                row_index=imported_rows,
                                question=None
                            )
                            response.set_value(str(value))
                        except Exception as e:
                            errors.append(f"Qator {imported_rows + 1}, ustun {column.title}: {str(e)}")

            imported_rows += 1

        # Progress yangilash
        assignment.update_progress()

        return JsonResponse({
            'success': True,
            'imported_rows': imported_rows,
            'progress': assignment.progress,
            'errors': errors if errors else None,
            'message': f"{imported_rows} ta qator muvaffaqiyatli import qilindi"
        })

    except TaskAssignment.DoesNotExist:
        return JsonResponse({
            'success': False,
            'error': 'Vazifa topilmadi'
        }, status=404)
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)


@login_required
def task_excel_template(request, pk):
    """
    Vazifa uchun Excel shablon yuklab olish
    Yetakchi shu shablonni to'ldirib, import qila oladi
    """
    from django.http import HttpResponse

    try:
        assignment = TaskAssignment.objects.select_related('task').get(
            pk=pk,
            leader=request.user
        )

        task = assignment.task

        if task.type != Task.Type.TABLE:
            return JsonResponse({
                'success': False,
                'error': 'Faqat jadval turidagi vazifalar uchun'
            })

        # Excel yaratish
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Ma'lumotlar"

        # Ustunlar
        columns = task.columns.order_by('order')

        # Sarlavha
        header_fill = openpyxl.styles.PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = openpyxl.styles.Font(bold=True, color="FFFFFF")

        for col_idx, column in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=column.title)
            cell.fill = header_fill
            cell.font = header_font

            # Ustun kengligi
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = max(len(column.title) + 5, 15)

        # Namuna qator (2-qator)
        for col_idx, column in enumerate(columns, 1):
            placeholder = ""
            if column.data_type == 'text':
                placeholder = "Matn kiriting"
            elif column.data_type == 'number':
                placeholder = "123"
            elif column.data_type == 'date':
                placeholder = "2025-01-31"
            elif column.data_type == 'phone':
                placeholder = "+998901234567"
            elif column.placeholder:
                placeholder = column.placeholder

            cell = ws.cell(row=2, column=col_idx, value=placeholder)
            cell.font = openpyxl.styles.Font(italic=True, color="888888")

        # Response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{task.title}_shablon.xlsx"'
        wb.save(response)

        return response

    except TaskAssignment.DoesNotExist:
        return JsonResponse({
            'success': False,
            'error': 'Vazifa topilmadi'
        }, status=404)


@login_required
@require_http_methods(["POST"])
def task_delete_row(request, pk):
    """Jadvaldan qatorni o'chirish"""
    try:
        assignment = TaskAssignment.objects.select_related('task').get(
            pk=pk,
            leader=request.user
        )

        if not assignment.can_edit:
            return JsonResponse({
                'success': False,
                'error': 'Tahrirlash mumkin emas'
            })

        data = json.loads(request.body)
        row_index = data.get('row_index')

        if row_index is None:
            return JsonResponse({
                'success': False,
                'error': 'Qator raqami ko\'rsatilmagan'
            })

        # Qatorni o'chirish
        deleted_count, _ = assignment.responses.filter(
            column__isnull=False,
            row_index=row_index
        ).delete()

        # Keyingi qatorlarni qayta tartiblab qo'yish
        responses_to_update = assignment.responses.filter(
            column__isnull=False,
            row_index__gt=row_index
        ).order_by('row_index')

        for response in responses_to_update:
            response.row_index -= 1
            response.save(update_fields=['row_index'])

        # Progress yangilash
        assignment.update_progress()

        return JsonResponse({
            'success': True,
            'deleted_count': deleted_count,
            'progress': assignment.progress
        })

    except TaskAssignment.DoesNotExist:
        return JsonResponse({
            'success': False,
            'error': 'Vazifa topilmadi'
        }, status=404)
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)


@login_required
def get_notifications(request):
    """Bildirishnomalar (AJAX)"""
    notifications = request.user.notifications.filter(is_read=False)[:10]

    data = []
    for notif in notifications:
        data.append({
            'id': str(notif.pk),
            'type': notif.type,
            'title': notif.title,
            'message': notif.message[:100],
            'link': notif.link,
            'created_at': notif.created_at.isoformat()
        })

    return JsonResponse({
        'count': request.user.notifications.filter(is_read=False).count(),
        'notifications': data
    })
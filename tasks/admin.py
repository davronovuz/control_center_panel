from django.contrib import admin
from django.utils.translation import gettext_lazy as _
from django.utils.html import format_html
from django.utils import timezone
from django.http import HttpResponse
from django.urls import reverse
from django.db.models import Count, Q
import json
import openpyxl

from .models import TaskTemplate, Task, TaskColumn, TaskQuestion, TaskAssignment, TaskResponse, TaskHistory


# ============================================================
# INLINES
# ============================================================

class TaskColumnInline(admin.TabularInline):
    model = TaskColumn
    extra = 1
    fields = ['order', 'title', 'data_type', 'required', 'width', 'choices']
    ordering = ['order']

    def get_queryset(self, request):
        return super().get_queryset(request).order_by('order')


class TaskQuestionInline(admin.TabularInline):
    model = TaskQuestion
    extra = 1
    fields = ['order', 'text', 'answer_type', 'required', 'choices']
    ordering = ['order']

    def get_queryset(self, request):
        return super().get_queryset(request).order_by('order')


class TaskAssignmentInline(admin.TabularInline):
    model = TaskAssignment
    extra = 0
    fields = ['leader', 'status', 'progress', 'submitted_at', 'approved_by']
    readonly_fields = ['progress', 'submitted_at', 'approved_by']
    ordering = ['-created_at']
    autocomplete_fields = ['leader']

    def get_queryset(self, request):
        return super().get_queryset(request).select_related('leader', 'approved_by')


class TaskHistoryInline(admin.TabularInline):
    model = TaskHistory
    extra = 0
    fields = ['action', 'actor', 'description', 'created_at']
    readonly_fields = ['action', 'actor', 'description', 'created_at']
    ordering = ['-created_at']
    max_num = 20

    def has_add_permission(self, request, obj=None):
        return False

    def has_delete_permission(self, request, obj=None):
        return False


# ============================================================
# TASK TEMPLATE ADMIN
# ============================================================

@admin.register(TaskTemplate)
class TaskTemplateAdmin(admin.ModelAdmin):
    list_display = ['name', 'type_badge', 'is_active', 'is_system', 'created_by', 'created_at']
    list_filter = ['type', 'is_active', 'is_system']
    search_fields = ['name', 'description']
    ordering = ['name']

    readonly_fields = ['created_at', 'updated_at']

    fieldsets = (
        (None, {
            'fields': ('name', 'description', 'type')
        }),
        (_('Struktura'), {
            'fields': ('structure', 'settings'),
            'classes': ('collapse',)
        }),
        (_('Holat'), {
            'fields': ('is_active', 'is_system')
        }),
        (_('Meta'), {
            'fields': ('created_by', 'created_at', 'updated_at'),
            'classes': ('collapse',)
        }),
    )

    autocomplete_fields = ['created_by']

    def type_badge(self, obj):
        colors = {
            'table': '#0d6efd',
            'survey': '#6f42c1',
            'report': '#fd7e14',
            'file': '#20c997',
            'mixed': '#6c757d'
        }
        icons = {
            'table': '📊',
            'survey': '📋',
            'report': '📝',
            'file': '📁',
            'mixed': '🔀'
        }
        color = colors.get(obj.type, '#6c757d')
        icon = icons.get(obj.type, '📄')
        return format_html(
            '<span style="background:{}; color:white; padding:3px 10px; '
            'border-radius:10px; font-size:11px;">{} {}</span>',
            color, icon, obj.get_type_display()
        )

    type_badge.short_description = _("Turi")

    def save_model(self, request, obj, form, change):
        if not change:
            obj.created_by = request.user
        super().save_model(request, obj, form, change)


# ============================================================
# TASK ADMIN
# ============================================================

@admin.register(Task)
class TaskAdmin(admin.ModelAdmin):
    list_display = [
        'title',
        'type_badge',
        'status_badge',
        'priority_badge',
        'target_display',
        'deadline_display',
        'completion_display',
        'created_by',
        'created_at'
    ]
    list_filter = [
        'type', 'status', 'priority',
        'target_all', 'target_region', 'target_district',
        'requires_approval', 'is_recurring',
        'created_at', 'deadline'
    ]
    search_fields = ['title', 'description', 'created_by__username']
    ordering = ['-created_at']
    date_hierarchy = 'created_at'

    readonly_fields = [
        'total_assigned', 'total_started', 'total_submitted',
        'total_approved', 'total_rejected',
        'published_at', 'completed_at', 'created_at', 'updated_at'
    ]

    fieldsets = (
        (None, {
            'fields': ('title', 'description', 'instructions')
        }),
        (_('Turi va holat'), {
            'fields': ('type', 'status', 'priority', 'template')
        }),
        (_('Kimga'), {
            'fields': ('target_all', 'target_region', 'target_district', 'target_mahallas')
        }),
        (_('Muddat'), {
            'fields': ('start_date', 'deadline')
        }),
        (_('Jadval sozlamalari'), {
            'fields': ('allow_multiple_rows', 'min_rows', 'max_rows'),
            'classes': ('collapse',)
        }),
        (_('Fayl sozlamalari'), {
            'fields': ('allowed_extensions', 'max_file_size', 'max_files'),
            'classes': ('collapse',)
        }),
        (_('Jarayon sozlamalari'), {
            'fields': ('requires_approval', 'allow_edit_after_submit', 'auto_save'),
            'classes': ('collapse',)
        }),
        (_('Eslatmalar'), {
            'fields': ('reminder_enabled', 'reminder_days'),
            'classes': ('collapse',)
        }),
        (_('Takrorlanish'), {
            'fields': ('is_recurring', 'recurring_type', 'recurring_end_date', 'parent_task'),
            'classes': ('collapse',)
        }),
        (_('Statistika'), {
            'fields': (
                'total_assigned', 'total_started', 'total_submitted',
                'total_approved', 'total_rejected'
            ),
            'classes': ('collapse',)
        }),
        (_('Meta'), {
            'fields': ('created_by', 'published_at', 'completed_at', 'created_at', 'updated_at'),
            'classes': ('collapse',)
        }),
    )

    filter_horizontal = ['target_mahallas']
    autocomplete_fields = ['target_region', 'target_district', 'template', 'parent_task', 'created_by']
    inlines = [TaskColumnInline, TaskQuestionInline, TaskAssignmentInline, TaskHistoryInline]

    actions = [
        'publish_tasks',
        'complete_tasks',
        'cancel_tasks',
        'duplicate_tasks',
        'send_reminders',
        'update_statistics',
        'export_results_excel'
    ]

    def type_badge(self, obj):
        colors = {
            'table': '#0d6efd',
            'survey': '#6f42c1',
            'report': '#fd7e14',
            'file': '#20c997',
            'mixed': '#6c757d'
        }
        icons = {
            'table': '📊',
            'survey': '📋',
            'report': '📝',
            'file': '📁',
            'mixed': '🔀'
        }
        color = colors.get(obj.type, '#6c757d')
        icon = icons.get(obj.type, '📄')
        return format_html(
            '<span style="background:{}; color:white; padding:3px 10px; '
            'border-radius:10px; font-size:11px;">{} {}</span>',
            color, icon, obj.get_type_display()
        )

    type_badge.short_description = _("Turi")

    def status_badge(self, obj):
        colors = {
            'draft': '#6c757d',
            'active': '#198754',
            'paused': '#ffc107',
            'completed': '#0d6efd',
            'cancelled': '#dc3545',
            'archived': '#495057'
        }
        color = colors.get(obj.status, '#6c757d')
        return format_html(
            '<span style="background:{}; color:white; padding:3px 10px; '
            'border-radius:10px; font-size:11px;">{}</span>',
            color, obj.get_status_display()
        )

    status_badge.short_description = _("Holat")

    def priority_badge(self, obj):
        colors = {
            'low': '#198754',
            'medium': '#0d6efd',
            'high': '#ffc107',
            'urgent': '#dc3545'
        }
        icons = {
            'low': '▽',
            'medium': '◇',
            'high': '△',
            'urgent': '⚠'
        }
        color = colors.get(obj.priority, '#6c757d')
        icon = icons.get(obj.priority, '◇')
        return format_html(
            '<span style="color:{}; font-weight:bold;">{} {}</span>',
            color, icon, obj.get_priority_display()
        )

    priority_badge.short_description = _("Muhimlik")

    def target_display(self, obj):
        if obj.target_all:
            return format_html('<span style="color: green;">🌍 Hammaga</span>')

        if obj.target_mahallas.exists():
            count = obj.target_mahallas.count()
            return format_html(
                '<span title="{}">🏘 {} ta mahalla</span>',
                ', '.join(m.name for m in obj.target_mahallas.all()[:5]),
                count
            )

        if obj.target_district:
            return format_html('🏢 {}', obj.target_district.name)

        if obj.target_region:
            return format_html('📍 {}', obj.target_region.name)

        return format_html('<span style="color: red;">⚠ Tanlanmagan</span>')

    target_display.short_description = _("Kimga")

    def deadline_display(self, obj):
        if obj.is_overdue:
            return format_html(
                '<span style="color: red; font-weight: bold;">'
                '⚠ {} (o\'tgan)</span>',
                obj.deadline.strftime('%d.%m.%Y %H:%M')
            )

        remaining = obj.time_remaining
        if remaining:
            if 'kun' in str(remaining) or 'soat' in str(remaining):
                color = 'orange' if 'soat' in str(remaining) else 'green'
            else:
                color = 'red'

            return format_html(
                '{}<br><small style="color: {};">{}</small>',
                obj.deadline.strftime('%d.%m.%Y %H:%M'),
                color,
                remaining
            )

        return obj.deadline.strftime('%d.%m.%Y %H:%M')

    deadline_display.short_description = _("Muddat")

    def completion_display(self, obj):
        if obj.total_assigned == 0:
            return format_html('<span style="color: gray;">—</span>')

        rate = obj.completion_rate

        if rate >= 80:
            color = '#198754'
        elif rate >= 50:
            color = '#ffc107'
        else:
            color = '#dc3545'

        completed = obj.total_approved if obj.requires_approval else obj.total_submitted

        return format_html(
            '<div style="width: 100px;">'
            '<div style="background: #e9ecef; border-radius: 4px; overflow: hidden;">'
            '<div style="width: {}%; background: {}; height: 8px;"></div>'
            '</div>'
            '<small>{}/{} ({}%)</small>'
            '</div>',
            rate, color, completed, obj.total_assigned, rate
        )

    completion_display.short_description = _("Bajarilish")

    def save_model(self, request, obj, form, change):
        if not change:
            obj.created_by = request.user
        super().save_model(request, obj, form, change)

        # Tarix
        if not change:
            TaskHistory.objects.create(
                task=obj,
                action=TaskHistory.Action.CREATED,
                actor=request.user,
                description=_("Vazifa yaratildi")
            )

    @admin.action(description=_("E'lon qilish"))
    def publish_tasks(self, request, queryset):
        success_count = 0
        error_messages = []

        for task in queryset.filter(status=Task.Status.DRAFT):
            success, message = task.publish(request.user)
            if success:
                success_count += 1
            else:
                error_messages.append(f"{task.title}: {message}")

        if success_count:
            self.message_user(request, f"{success_count} ta vazifa e'lon qilindi.")

        if error_messages:
            for msg in error_messages:
                self.message_user(request, msg, level='error')

    @admin.action(description=_("Yakunlash"))
    def complete_tasks(self, request, queryset):
        count = 0
        for task in queryset.filter(status=Task.Status.ACTIVE):
            task.complete()
            count += 1
        self.message_user(request, f"{count} ta vazifa yakunlandi.")

    @admin.action(description=_("Bekor qilish"))
    def cancel_tasks(self, request, queryset):
        count = queryset.exclude(
            status__in=[Task.Status.COMPLETED, Task.Status.CANCELLED]
        ).update(status=Task.Status.CANCELLED)
        self.message_user(request, f"{count} ta vazifa bekor qilindi.")

    @admin.action(description=_("Nusxalash"))
    def duplicate_tasks(self, request, queryset):
        count = 0
        for task in queryset:
            task.duplicate(request.user)
            count += 1
        self.message_user(request, f"{count} ta vazifa nusxalandi.")

    @admin.action(description=_("Eslatma yuborish"))
    def send_reminders(self, request, queryset):
        from accounts.models import Notification

        count = 0
        for task in queryset.filter(status=Task.Status.ACTIVE):
            pending_assignments = task.assignments.filter(
                status__in=[TaskAssignment.Status.PENDING, TaskAssignment.Status.VIEWED]
            )

            for assignment in pending_assignments:
                Notification.send(
                    user=assignment.leader,
                    type=Notification.Type.TASK_DEADLINE,
                    title=_("Vazifa eslatmasi"),
                    message=f"'{task.title}' vazifasi bajarilmagan. Muddat: {task.deadline.strftime('%d.%m.%Y %H:%M')}",
                    link=f'/leader/tasks/{task.pk}/',
                    priority='high'
                )
                count += 1

        self.message_user(request, f"{count} ta eslatma yuborildi.")

    @admin.action(description=_("Statistikani yangilash"))
    def update_statistics(self, request, queryset):
        for task in queryset:
            task.update_stats()
        self.message_user(request, f"{queryset.count()} ta vazifa statistikasi yangilandi.")

    @admin.action(description=_("Natijalarni Excel ga eksport"))
    def export_results_excel(self, request, queryset):
        if queryset.count() != 1:
            self.message_user(request, "Faqat 1 ta vazifa tanlang.", level='error')
            return

        task = queryset.first()

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Natijalar"

        # Sarlavha
        headers = ['№', 'Yetakchi', 'Mahalla', 'Tuman', 'Holat', 'Yuborilgan']

        if task.type == Task.Type.TABLE:
            for column in task.columns.order_by('order'):
                headers.append(column.title)
        elif task.type == Task.Type.SURVEY:
            for question in task.questions.order_by('order'):
                headers.append(question.text[:50])

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = openpyxl.styles.Font(bold=True)

        # Ma'lumotlar
        row_num = 2
        for assignment in task.assignments.select_related('leader', 'leader__mahalla', 'leader__district'):
            ws.cell(row=row_num, column=1, value=row_num - 1)
            ws.cell(row=row_num, column=2, value=assignment.leader.get_full_name())
            ws.cell(row=row_num, column=3, value=str(assignment.leader.mahalla) if assignment.leader.mahalla else '')
            ws.cell(row=row_num, column=4, value=str(assignment.leader.district) if assignment.leader.district else '')
            ws.cell(row=row_num, column=5, value=assignment.get_status_display())
            ws.cell(row=row_num, column=6,
                    value=assignment.submitted_at.strftime('%d.%m.%Y %H:%M') if assignment.submitted_at else '')

            col_num = 7
            if task.type == Task.Type.TABLE:
                for column in task.columns.order_by('order'):
                    response = assignment.responses.filter(column=column).first()
                    value = response.display_value if response else ''
                    ws.cell(row=row_num, column=col_num, value=value)
                    col_num += 1
            elif task.type == Task.Type.SURVEY:
                for question in task.questions.order_by('order'):
                    response = assignment.responses.filter(question=question).first()
                    value = response.display_value if response else ''
                    ws.cell(row=row_num, column=col_num, value=value)
                    col_num += 1

            row_num += 1

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{task.title}_natijalar.xlsx"'
        wb.save(response)
        return response


# ============================================================
# TASK COLUMN ADMIN
# ============================================================

@admin.register(TaskColumn)
class TaskColumnAdmin(admin.ModelAdmin):
    list_display = ['task', 'order', 'title', 'data_type_badge', 'required_badge', 'width']
    list_filter = ['data_type', 'required', 'task']
    search_fields = ['title', 'task__title']
    ordering = ['task', 'order']

    autocomplete_fields = ['task']

    def data_type_badge(self, obj):
        colors = {
            'text': '#6c757d',
            'textarea': '#6c757d',
            'number': '#0d6efd',
            'decimal': '#0d6efd',
            'date': '#6f42c1',
            'datetime': '#6f42c1',
            'time': '#6f42c1',
            'choice': '#20c997',
            'multiple': '#20c997',
            'boolean': '#fd7e14',
            'phone': '#198754',
            'email': '#198754',
            'url': '#198754',
            'file': '#dc3545',
            'image': '#dc3545'
        }
        color = colors.get(obj.data_type, '#6c757d')
        return format_html(
            '<span style="background:{}; color:white; padding:2px 8px; '
            'border-radius:8px; font-size:10px;">{}</span>',
            color, obj.get_data_type_display()
        )

    data_type_badge.short_description = _("Turi")

    def required_badge(self, obj):
        if obj.required:
            return format_html('<span style="color: red;">* Majburiy</span>')
        return format_html('<span style="color: gray;">Ixtiyoriy</span>')

    required_badge.short_description = _("Majburiy")


# ============================================================
# TASK QUESTION ADMIN
# ============================================================

@admin.register(TaskQuestion)
class TaskQuestionAdmin(admin.ModelAdmin):
    list_display = ['task', 'order', 'text_short', 'answer_type_badge', 'required_badge']
    list_filter = ['answer_type', 'required', 'task']
    search_fields = ['text', 'task__title']
    ordering = ['task', 'order']

    autocomplete_fields = ['task', 'depends_on']

    def text_short(self, obj):
        if len(obj.text) > 50:
            return obj.text[:50] + '...'
        return obj.text

    text_short.short_description = _("Savol")

    def answer_type_badge(self, obj):
        colors = {
            'text': '#6c757d',
            'textarea': '#6c757d',
            'number': '#0d6efd',
            'date': '#6f42c1',
            'choice': '#20c997',
            'multiple': '#20c997',
            'boolean': '#fd7e14',
            'rating': '#ffc107',
            'file': '#dc3545'
        }
        color = colors.get(obj.answer_type, '#6c757d')
        return format_html(
            '<span style="background:{}; color:white; padding:2px 8px; '
            'border-radius:8px; font-size:10px;">{}</span>',
            color, obj.get_answer_type_display()
        )

    answer_type_badge.short_description = _("Turi")

    def required_badge(self, obj):
        if obj.required:
            return format_html('<span style="color: red;">* Majburiy</span>')
        return format_html('<span style="color: gray;">Ixtiyoriy</span>')

    required_badge.short_description = _("Majburiy")


# ============================================================
# TASK ASSIGNMENT ADMIN
# ============================================================

@admin.register(TaskAssignment)
class TaskAssignmentAdmin(admin.ModelAdmin):
    list_display = [
        'task',
        'leader',
        'leader_location',
        'status_badge',
        'progress_display',
        'submitted_at',
        'approved_by',
        'created_at'
    ]
    list_filter = ['status', 'task', 'task__type', 'created_at', 'submitted_at']
    search_fields = ['task__title', 'leader__username', 'leader__first_name', 'leader__last_name']
    ordering = ['-created_at']
    date_hierarchy = 'created_at'

    readonly_fields = [
        'viewed_at', 'started_at', 'submitted_at',
        'approved_at', 'rejected_at', 'progress',
        'created_at', 'updated_at'
    ]

    fieldsets = (
        (None, {
            'fields': ('task', 'leader', 'status')
        }),
        (_('Vaqtlar'), {
            'fields': ('viewed_at', 'started_at', 'submitted_at', 'approved_at', 'rejected_at')
        }),
        (_('Tasdiqlash'), {
            'fields': ('approved_by', 'rejection_reason', 'admin_notes')
        }),
        (_('Jarayon'), {
            'fields': ('progress',)
        }),
        (_('Meta'), {
            'fields': ('created_at', 'updated_at'),
            'classes': ('collapse',)
        }),
    )

    autocomplete_fields = ['task', 'leader', 'approved_by']

    actions = ['approve_assignments', 'reject_assignments', 'reset_assignments']

    def leader_location(self, obj):
        if obj.leader.mahalla:
            return format_html(
                '<small>{}</small>',
                obj.leader.mahalla.name
            )
        return format_html('<span style="color: gray;">—</span>')

    leader_location.short_description = _("Mahalla")

    def status_badge(self, obj):
        colors = {
            'pending': '#6c757d',
            'viewed': '#17a2b8',
            'in_progress': '#ffc107',
            'submitted': '#0d6efd',
            'approved': '#198754',
            'rejected': '#dc3545',
            'overdue': '#dc3545'
        }

        status = obj.status
        if obj.is_overdue and status not in ['submitted', 'approved']:
            status = 'overdue'
            label = _("Muddati o'tdi")
        else:
            label = obj.get_status_display()

        color = colors.get(status, '#6c757d')
        return format_html(
            '<span style="background:{}; color:white; padding:3px 10px; '
            'border-radius:10px; font-size:11px;">{}</span>',
            color, label
        )

    status_badge.short_description = _("Holat")

    def progress_display(self, obj):
        progress = obj.progress

        if progress >= 100:
            color = '#198754'
        elif progress >= 50:
            color = '#ffc107'
        else:
            color = '#dc3545'

        return format_html(
            '<div style="width: 80px;">'
            '<div style="background: #e9ecef; border-radius: 4px; overflow: hidden;">'
            '<div style="width: {}%; background: {}; height: 6px;"></div>'
            '</div>'
            '<small>{}%</small>'
            '</div>',
            progress, color, progress
        )

    progress_display.short_description = _("Jarayon")

    @admin.action(description=_("Tasdiqlash"))
    def approve_assignments(self, request, queryset):
        count = 0
        for assignment in queryset.filter(status=TaskAssignment.Status.SUBMITTED):
            assignment.approve(request.user)
            count += 1
        self.message_user(request, f"{count} ta vazifa tasdiqlandi.")

    @admin.action(description=_("Rad etish"))
    def reject_assignments(self, request, queryset):
        count = queryset.filter(status=TaskAssignment.Status.SUBMITTED).count()
        queryset.filter(status=TaskAssignment.Status.SUBMITTED).update(
            status=TaskAssignment.Status.REJECTED,
            rejected_at=timezone.now(),
            approved_by=request.user
        )
        self.message_user(request, f"{count} ta vazifa rad etildi.")

    @admin.action(description=_("Qayta boshlash"))
    def reset_assignments(self, request, queryset):
        count = queryset.update(
            status=TaskAssignment.Status.PENDING,
            viewed_at=None,
            started_at=None,
            submitted_at=None,
            approved_at=None,
            rejected_at=None,
            approved_by=None,
            progress=0
        )
        self.message_user(request, f"{count} ta tayinlash qayta boshlandi.")


# ============================================================
# TASK RESPONSE ADMIN
# ============================================================

@admin.register(TaskResponse)
class TaskResponseAdmin(admin.ModelAdmin):
    list_display = ['assignment', 'source_display', 'row_index', 'value_display', 'is_valid_badge', 'updated_at']
    list_filter = ['is_valid', 'assignment__task', 'created_at']
    search_fields = ['assignment__leader__username', 'assignment__task__title']
    ordering = ['-updated_at']

    readonly_fields = ['is_valid', 'validation_errors', 'created_at', 'updated_at']

    autocomplete_fields = ['assignment', 'column', 'question']

    def source_display(self, obj):
        if obj.column:
            return format_html('<span style="color: #0d6efd;">📊 {}</span>', obj.column.title)
        if obj.question:
            text = obj.question.text[:30] + '...' if len(obj.question.text) > 30 else obj.question.text
            return format_html('<span style="color: #6f42c1;">📋 {}</span>', text)
        return '—'

    source_display.short_description = _("Ustun/Savol")

    def value_display(self, obj):
        value = obj.display_value
        if len(str(value)) > 50:
            return str(value)[:50] + '...'
        return value

    value_display.short_description = _("Qiymat")

    def is_valid_badge(self, obj):
        if obj.is_valid:
            return format_html('<span style="color: green;">✓</span>')
        errors = ', '.join(obj.validation_errors) if obj.validation_errors else ''
        return format_html(
            '<span style="color: red;" title="{}">✗</span>',
            errors
        )

    is_valid_badge.short_description = _("✓")


# ============================================================
# TASK HISTORY ADMIN
# ============================================================

@admin.register(TaskHistory)
class TaskHistoryAdmin(admin.ModelAdmin):
    list_display = ['task', 'action_badge', 'actor', 'description_short', 'ip_address', 'created_at']
    list_filter = ['action', 'created_at']
    search_fields = ['task__title', 'actor__username', 'description']
    ordering = ['-created_at']
    date_hierarchy = 'created_at'

    readonly_fields = [
        'task', 'assignment', 'action', 'actor',
        'description', 'changes', 'ip_address', 'user_agent', 'created_at'
    ]

    def action_badge(self, obj):
        colors = {
            'created': '#0d6efd',
            'updated': '#17a2b8',
            'published': '#198754',
            'assigned': '#6f42c1',
            'viewed': '#6c757d',
            'started': '#ffc107',
            'saved': '#20c997',
            'submitted': '#0d6efd',
            'approved': '#198754',
            'rejected': '#dc3545',
            'completed': '#198754',
            'cancelled': '#dc3545',
            'reminder': '#fd7e14'
        }
        color = colors.get(obj.action, '#6c757d')
        return format_html(
            '<span style="background:{}; color:white; padding:2px 8px; '
            'border-radius:8px; font-size:10px;">{} {}</span>',
            color, obj.action_icon, obj.get_action_display()
        )

    action_badge.short_description = _("Harakat")

    def description_short(self, obj):
        if len(obj.description) > 50:
            return obj.description[:50] + '...'
        return obj.description or '—'

    description_short.short_description = _("Tavsif")

    def has_add_permission(self, request):
        return False

    def has_change_permission(self, request, obj=None):
        return False

    def has_delete_permission(self, request, obj=None):
        return request.user.is_superuser
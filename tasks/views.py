from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.core.paginator import Paginator
from django.http import HttpResponse
from django.db.models import Q, Count
from django.utils import timezone
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from .models import Task, TaskColumn, TaskQuestion, TaskAssignment, TaskHistory, TaskResponse


# ==============================================================================
# YORDAMCHI FUNKSIYALAR
# ==============================================================================
def get_list_secure(request, key):
    """
    HTML formadan kelgan ma'lumotlarni 'name[]' va 'name' formatlarida tekshirib,
    har qanday holatda ham to'liq ro'yxatni qaytaradi.
    """
    # 1. 'name[]' formatida (agar HTML da qavs bo'lsa)
    values = request.POST.getlist(f"{key}[]")

    # 2. Agar bo'sh bo'lsa, 'name' formatida tekshiramiz (agar HTML da qavs bo'lmasa)
    if not values:
        values = request.POST.getlist(key)

    return values


def group_responses_by_row(assignment):
    """
    Assignment ichidagi javoblarni qatorlar (row_index) bo'yicha guruhlaydi.
    Natija: [
        {'row_index': 0, 'answers': {'col_id': 'val1'}},
        {'row_index': 1, 'answers': {'col_id': 'val2'}},
    ]
    """
    # Barcha javoblarni olamiz
    responses = assignment.responses.select_related('column', 'question').all()

    # Guruhlash uchun vaqtinchalik lug'at: { row_index: { col_id: value } }
    grouped = {}

    for resp in responses:
        # Agar row_index bo'lmasa, 0 deb olamiz
        idx = resp.row_index if resp.row_index is not None else 0

        if idx not in grouped:
            grouped[idx] = {}

        if resp.column:
            grouped[idx][str(resp.column.pk)] = resp.display_value
        elif resp.question:
            grouped[idx][str(resp.question.pk)] = resp.display_value
        else:
            grouped[idx]['report'] = resp.value_text

    # Agar javoblar bo'lmasa, lekin assignment bor bo'lsa (bo'sh qator)
    if not grouped and assignment.status in ['submitted', 'approved', 'rejected']:
        return [{'row_index': 0, 'answers': {}}]

    # Ro'yxatga aylantiramiz va tartiblaymiz
    result_rows = []
    for idx in sorted(grouped.keys()):
        result_rows.append({
            'row_index': idx,
            'answers': grouped[idx]
        })

    return result_rows


# ==============================================================================
# VIEW FUNKSIYALAR
# ==============================================================================

@login_required
def task_list(request):
    """Vazifalar ro'yxati"""
    tasks = Task.objects.select_related('created_by', 'target_region', 'target_district').all().order_by('-created_at')

    # Filterlar
    status = request.GET.get('status')
    priority = request.GET.get('priority')
    task_type = request.GET.get('type')
    search = request.GET.get('search')

    if status:
        tasks = tasks.filter(status=status)
    if priority:
        tasks = tasks.filter(priority=priority)
    if task_type:
        tasks = tasks.filter(type=task_type)
    if search:
        tasks = tasks.filter(
            Q(title__icontains=search) | Q(description__icontains=search)
        )

    # Pagination
    paginator = Paginator(tasks, 20)
    page = request.GET.get('page')
    tasks = paginator.get_page(page)

    context = {
        'tasks': tasks,
        'status_choices': Task.Status.choices,
        'priority_choices': Task.Priority.choices,
        'type_choices': Task.Type.choices,
        'current_filters': {
            'status': status,
            'priority': priority,
            'type': task_type,
            'search': search,
        }
    }
    return render(request, 'tasks/task_list.html', context)


@login_required
def task_detail(request, pk):
    """Vazifa tafsilotlari va statistikasi"""
    task = get_object_or_404(Task.objects.select_related('created_by'), pk=pk)

    columns = task.columns.all().order_by('order')
    questions = task.questions.all().order_by('order')

    assignments = task.assignments.select_related('leader', 'leader__mahalla', 'leader__district').all()

    # Pagination for assignments
    paginator = Paginator(assignments, 20)
    page = request.GET.get('page')
    assignments = paginator.get_page(page)

    # Statistics
    stats = {
        'total': task.total_assigned,
        'pending': task.assignments.filter(status=TaskAssignment.Status.PENDING).count(),
        'viewed': task.assignments.filter(status=TaskAssignment.Status.VIEWED).count(),
        'in_progress': task.assignments.filter(status=TaskAssignment.Status.IN_PROGRESS).count(),
        'submitted': task.assignments.filter(status=TaskAssignment.Status.SUBMITTED).count(),
        'approved': task.assignments.filter(status=TaskAssignment.Status.APPROVED).count(),
        'rejected': task.assignments.filter(status=TaskAssignment.Status.REJECTED).count(),
    }

    context = {
        'task': task,
        'columns': columns,
        'questions': questions,
        'assignments': assignments,
        'stats': stats,
    }
    return render(request, 'tasks/task_detail.html', context)


@login_required
def task_create(request):
    """Yangi vazifa yaratish (Konstruktor)"""
    from accounts.models import Region, District

    if request.method == 'POST':
        # --- 1. Asosiy ma'lumotlar ---
        title = request.POST.get('title')
        description = request.POST.get('description', '')
        instructions = request.POST.get('instructions', '')
        task_type = request.POST.get('type', 'table')
        priority = request.POST.get('priority', 'medium')
        deadline = request.POST.get('deadline')

        # Targeting
        target_all = request.POST.get('target_all') == 'on'
        target_region_id = request.POST.get('target_region')
        target_district_id = request.POST.get('target_district')

        # Settings
        requires_approval = request.POST.get('requires_approval') == 'on'
        allow_multiple_rows = request.POST.get('allow_multiple_rows') == 'on'

        # --- 2. Vazifani yaratish ---
        task = Task.objects.create(
            title=title,
            description=description,
            instructions=instructions,
            type=task_type,
            priority=priority,
            deadline=deadline if deadline else None,
            target_all=target_all,
            target_region_id=target_region_id if target_region_id else None,
            target_district_id=target_district_id if target_district_id else None,
            requires_approval=requires_approval,
            allow_multiple_rows=allow_multiple_rows,
            created_by=request.user,
            status=Task.Status.DRAFT  # Avval qoralama
        )

        # --- 3. Ustunlarni saqlash (TABLE) ---
        if task_type == 'table':
            # MUHIM: Xavfsiz funksiya orqali olish
            column_titles = get_list_secure(request, 'column_title')
            column_types = get_list_secure(request, 'column_type')

            for i, (col_title, col_type) in enumerate(zip(column_titles, column_types)):
                if col_title.strip():
                    TaskColumn.objects.create(
                        task=task,
                        title=col_title.strip(),
                        data_type=col_type,
                        order=i + 1
                    )

        # --- 4. Savollarni saqlash (SURVEY) ---
        elif task_type == 'survey':
            question_texts = get_list_secure(request, 'question_text')
            question_types = get_list_secure(request, 'question_type')

            for i, (q_text, q_type) in enumerate(zip(question_texts, question_types)):
                if q_text.strip():
                    TaskQuestion.objects.create(
                        task=task,
                        text=q_text.strip(),
                        answer_type=q_type,
                        order=i + 1
                    )

        # Tarix
        TaskHistory.objects.create(
            task=task,
            action=TaskHistory.Action.CREATED,
            actor=request.user,
            description="Vazifa konstruktori orqali yaratildi"
        )

        messages.success(request, "Vazifa muvaffaqiyatli yaratildi!")
        return redirect('tasks:task_detail', pk=task.pk)

    context = {
        'regions': Region.objects.filter(is_active=True),
        'type_choices': Task.Type.choices,
        'priority_choices': Task.Priority.choices,
        'column_type_choices': TaskColumn.DataType.choices,
        'question_type_choices': TaskQuestion.AnswerType.choices,
    }
    return render(request, 'tasks/task_form.html', context)


@login_required
def task_edit(request, pk):
    """Vazifani tahrirlash"""
    from accounts.models import Region, District

    task = get_object_or_404(Task, pk=pk)

    if task.status != Task.Status.DRAFT:
        messages.error(request, "Faqat qoralama (Draft) holatidagi vazifani tahrirlash mumkin!")
        return redirect('tasks:task_detail', pk=pk)

    if request.method == 'POST':
        task.title = request.POST.get('title')
        task.description = request.POST.get('description', '')
        task.instructions = request.POST.get('instructions', '')
        task.priority = request.POST.get('priority', 'medium')
        task.deadline = request.POST.get('deadline')

        # Checkboxlar
        task.target_all = request.POST.get('target_all') == 'on'
        task.requires_approval = request.POST.get('requires_approval') == 'on'
        task.allow_multiple_rows = request.POST.get('allow_multiple_rows') == 'on'

        # Selectlar
        r_id = request.POST.get('target_region')
        d_id = request.POST.get('target_district')
        task.target_region_id = r_id if r_id else None
        task.target_district_id = d_id if d_id else None

        task.save()

        # --- Ustunlarni yangilash ---
        if task.type == 'table':
            # Eski ustunlarni o'chirib, yangilarini yozamiz
            task.columns.all().delete()

            # Yana o'sha xavfsiz funksiya
            titles = get_list_secure(request, 'column_title')
            types = get_list_secure(request, 'column_type')

            for i, (t, dt) in enumerate(zip(titles, types)):
                if t.strip():
                    TaskColumn.objects.create(task=task, title=t.strip(), data_type=dt, order=i + 1)

        # --- Savollarni yangilash ---
        elif task.type == 'survey':
            task.questions.all().delete()

            q_texts = get_list_secure(request, 'question_text')
            q_types = get_list_secure(request, 'question_type')

            for i, (qt, qtp) in enumerate(zip(q_texts, q_types)):
                if qt.strip():
                    TaskQuestion.objects.create(task=task, text=qt.strip(), answer_type=qtp, order=i + 1)

        messages.success(request, "Vazifa muvaffaqiyatli yangilandi!")
        return redirect('tasks:task_detail', pk=task.pk)

    context = {
        'task': task,
        'columns': task.columns.order_by('order'),
        'questions': task.questions.order_by('order'),
        'regions': Region.objects.filter(is_active=True),
        'districts': District.objects.filter(region=task.target_region) if task.target_region else [],
        'type_choices': Task.Type.choices,
        'priority_choices': Task.Priority.choices,
        'column_type_choices': TaskColumn.DataType.choices,
        'question_type_choices': TaskQuestion.AnswerType.choices,
    }
    return render(request, 'tasks/task_form.html', context)


@login_required
def task_delete(request, pk):
    """Vazifani o'chirish"""
    task = get_object_or_404(Task, pk=pk)

    if request.method == 'POST':
        task.delete()
        messages.success(request, "Vazifa o'chirildi!")
        return redirect('tasks:task_list')

    return render(request, 'tasks/task_delete.html', {'task': task})


@login_required
def task_publish(request, pk):
    """Vazifani e'lon qilish"""
    task = get_object_or_404(Task, pk=pk)

    if request.method == 'POST':
        # Agar modelda maxsus publish metodi bo'lsa, uni ishlatamiz
        if hasattr(task, 'publish'):
            success, message = task.publish(request.user)
            if success:
                messages.success(request, message)
            else:
                messages.error(request, message)
        else:
            # Oddiy holatda statusni o'zgartiramiz
            task.status = Task.Status.PUBLISHED
            task.save()
            messages.success(request, "Vazifa e'lon qilindi")

        return redirect('tasks:task_detail', pk=pk)

    # Preview uchun ma'lumotlar
    target_leaders = []
    if hasattr(task, 'get_target_leaders'):
        target_leaders = task.get_target_leaders()[:20]
        target_leaders_count = task.get_target_leaders().count()
    else:
        target_leaders_count = 0

    context = {
        'task': task,
        'target_leaders': target_leaders,
        'target_leaders_count': target_leaders_count,
    }
    return render(request, 'tasks/task_publish.html', context)


@login_required
def task_results(request, pk):
    """Vazifa natijalari"""
    task = get_object_or_404(Task, pk=pk)

    # Faqat yuborilgan (submitted/approved/rejected) assignmentlar
    assignments = task.assignments.filter(
        status__in=['submitted', 'approved', 'rejected']
    ).select_related('leader', 'leader__mahalla', 'leader__district').prefetch_related('responses')

    columns = list(task.columns.order_by('order'))
    questions = list(task.questions.order_by('order'))

    # Natijalarni "yassilash" (Flattening)
    # 1 Assignment = N ta qator (Row) bo'lishi mumkin
    flat_results = []

    for assignment in assignments:
        # Har bir assignmentni qatorlarga ajratamiz (Helper funksiya yordamida)
        rows = group_responses_by_row(assignment)

        for row in rows:
            flat_results.append({
                'assignment': assignment,  # Asl assignment obyekti
                'leader': assignment.leader,  # Yetakchi ma'lumotlari
                'status': assignment.status,
                'submitted_at': assignment.submitted_at,
                'row_index': row['row_index'],  # Qator raqami (0, 1, 2...)
                'answers': row['answers'],  # Shu qatorga tegishli javoblar {col_id: value}
            })

    context = {
        'task': task,
        'columns': columns,
        'questions': questions,
        'results': flat_results,  # Endi bu ro'yxatda barcha qatorlar bor
    }
    return render(request, 'tasks/task_results.html', context)


@login_required
def task_export(request, pk):
    """Excelga eksport"""
    task = get_object_or_404(Task, pk=pk)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Natijalar"

    # --- Header Style ---
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    center_align = Alignment(horizontal='center', vertical='center')
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))

    # Sarlavhalar
    headers = ['№', 'Yetakchi', 'Mahalla', 'Tuman', 'Telefon']

    if task.type == Task.Type.TABLE:
        for column in task.columns.order_by('order'):
            headers.append(column.title)
    elif task.type == Task.Type.SURVEY:
        for question in task.questions.order_by('order'):
            headers.append(question.text[:50])

    headers.extend(['Holat', 'Vaqt'])

    # Header yozish
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

    # Ma'lumotlarni yig'ish
    assignments = task.assignments.filter(
        status__in=['submitted', 'approved', 'rejected']
    ).select_related(
        'leader', 'leader__mahalla', 'leader__district'
    ).prefetch_related('responses')

    current_row = 2
    counter = 1

    for assignment in assignments:
        # Assignmentni qatorlarga ajratamiz (1 assignment = N rows)
        rows = group_responses_by_row(assignment)

        for row_data in rows:
            # 1. Meta ma'lumotlar
            ws.cell(row=current_row, column=1, value=counter).border = thin_border
            ws.cell(row=current_row, column=2, value=assignment.leader.get_full_name()).border = thin_border
            ws.cell(row=current_row, column=3, value=str(assignment.leader.mahalla or '')).border = thin_border
            ws.cell(row=current_row, column=4, value=str(assignment.leader.district or '')).border = thin_border
            ws.cell(row=current_row, column=5, value=assignment.leader.phone or '').border = thin_border

            # 2. Javoblar
            col_idx = 6
            answers = row_data['answers']  # Faqat shu qator javoblari

            if task.type == Task.Type.TABLE:
                for column in task.columns.order_by('order'):
                    val = answers.get(str(column.pk), '')
                    ws.cell(row=current_row, column=col_idx, value=val).border = thin_border
                    col_idx += 1
            elif task.type == Task.Type.SURVEY:
                for question in task.questions.order_by('order'):
                    val = answers.get(str(question.pk), '')
                    ws.cell(row=current_row, column=col_idx, value=val).border = thin_border
                    col_idx += 1

            # 3. Status
            ws.cell(row=current_row, column=col_idx, value=assignment.get_status_display()).border = thin_border

            # 4. Vaqt
            time_str = assignment.submitted_at.strftime('%d.%m.%Y %H:%M') if assignment.submitted_at else ''
            ws.cell(row=current_row, column=col_idx + 1, value=time_str).border = thin_border

            current_row += 1
            counter += 1

    # Ustun kengligini avtomatik moslash
    for column_cells in ws.columns:
        length = max(len(str(cell.value) or "") for cell in column_cells)
        ws.column_dimensions[column_cells[0].column_letter].width = min(length + 2, 50)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    # Fayl nomidagi kirill yoki maxsus belgilarni to'g'irlash uchun
    filename = f"task_results_{task.pk}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


import pandas as pd
import datetime
from django.shortcuts import render, redirect
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.utils import timezone
from django.db import transaction
from .models import Task, TaskHistory
from .forms import ExcelUploadForm


@login_required
def import_tasks_excel(request):
    """
    Excel fayldan vazifalarni professional darajada import qilish.
    Xatolarni ushlaydi va foydalanuvchiga batafsil hisobot beradi.
    """
    if request.method == 'POST':
        form = ExcelUploadForm(request.POST, request.FILES)
        if form.is_valid():
            excel_file = request.FILES['excel_file']

            try:
                # 1. Faylni o'qish (faqat birinchi varaqni)
                df = pd.read_excel(excel_file, engine='openpyxl')

                # 2. Ustun nomlarini standartlashtirish (probelni olish, kichik harf qilish)
                df.columns = [c.strip().lower() for c in df.columns]

                # Talab qilinadigan ustunlar (Exceldagi nomlar kichik harfda)
                required_columns = ['title']
                # Mavjud ustunlarni tekshirish
                missing_cols = [col for col in required_columns if col not in df.columns]

                if missing_cols:
                    messages.error(request, f"Faylda majburiy ustunlar yetishmayapti: {', '.join(missing_cols)}")
                    return redirect('tasks:import_excel')

                success_count = 0
                error_list = []

                # 3. Tranzaksiya ochamiz (xavfsizlik uchun)
                # Agar biror jiddiy xato bo'lsa, hammasini orqaga qaytarish imkoni bo'lishi kerak,
                # lekin biz bu yerda "skip error" taktikasini qo'llaymiz.

                for index, row in df.iterrows():
                    row_num = index + 2  # Excelda sarlavha 1-qator, ma'lumot 2-qatordan boshlanadi

                    try:
                        # --- Title (Majburiy) ---
                        title = row.get('title')
                        if pd.isna(title) or str(title).strip() == '':
                            error_list.append(f"Qator {row_num}: 'Title' bo'sh bo'lishi mumkin emas.")
                            continue  # Keyingi qatorga o'tamiz

                        # --- Description & Instructions ---
                        description = row.get('description', '')
                        instructions = row.get('instructions', '')

                        if pd.isna(description): description = ""
                        if pd.isna(instructions): instructions = ""

                        # --- Priority ---
                        priority_raw = str(row.get('priority', 'medium')).lower().strip()
                        priority_map = {
                            'yuqori': 'high', 'high': 'high', '3': 'high',
                            'orta': 'medium', 'o\'rta': 'medium', 'medium': 'medium', '2': 'medium',
                            'past': 'low', 'low': 'low', '1': 'low'
                        }
                        priority = priority_map.get(priority_raw, 'medium')

                        # --- Deadline (Eng nozik qism) ---
                        deadline_val = row.get('deadline')
                        final_deadline = None

                        if pd.notnull(deadline_val):
                            try:
                                # Agar allaqachon datetime bo'lsa
                                if isinstance(deadline_val, datetime.datetime):
                                    final_deadline = deadline_val
                                # Agar string bo'lsa (2024-12-31 yoki 31.12.2024)
                                elif isinstance(deadline_val, str):
                                    final_deadline = pd.to_datetime(deadline_val, dayfirst=True)

                                # Timezone aware qilish (Django ogohlantirish bermasligi uchun)
                                if final_deadline and timezone.is_naive(final_deadline):
                                    final_deadline = timezone.make_aware(final_deadline)
                            except:
                                # Sana xato bo'lsa, default +3 kun
                                final_deadline = timezone.now() + datetime.timedelta(days=3)
                        else:
                            final_deadline = timezone.now() + datetime.timedelta(days=3)

                        # --- Vazifani Yaratish ---
                        task = Task.objects.create(
                            title=str(title).strip(),
                            description=str(description).strip(),
                            instructions=str(instructions).strip(),
                            priority=priority,
                            deadline=final_deadline,
                            type='report',  # Default tur
                            target_all=True,  # Hozircha hammaga
                            requires_approval=True,
                            status=Task.Status.DRAFT,  # Qoralama
                            created_by=request.user
                        )

                        # Tarixga yozish
                        TaskHistory.objects.create(
                            task=task,
                            action=TaskHistory.Action.CREATED,
                            actor=request.user,
                            description=f"Excel orqali import qilindi (Qator {row_num})"
                        )

                        success_count += 1

                    except Exception as e:
                        error_list.append(f"Qator {row_num}: Tizim xatoligi - {str(e)}")

                # --- Natija ---
                if success_count > 0:
                    messages.success(request, f"✅ {success_count} ta vazifa muvaffaqiyatli yuklandi!")

                if error_list:
                    # Xatolarni ko'rsatish (maksimum 5 tasini, juda ko'payib ketmasligi uchun)
                    error_msg = "<br>".join(error_list[:5])
                    if len(error_list) > 5:
                        error_msg += f"<br>...va yana {len(error_list) - 5} ta xato."
                    messages.warning(request,
                                     f"⚠️ Ba'zi qatorlarda xatoliklar bo'ldi:<br>{error_msg}")  # 'safe' filtri kerak bo'ladi shablonda

                return redirect('tasks:task_list')

            except Exception as e:
                messages.error(request, f"Excel faylni o'qishda jiddiy xatolik: {str(e)}")
                return redirect('tasks:import_excel')

    else:
        form = ExcelUploadForm()

    return render(request, 'tasks/import_excel.html', {'form': form})
"""
Django Admin - Mahalla yetakchilarini JSON dan import qilish
Login/Parol generatsiya qilish va Excel yuklab olish

Bu faylni admin.py ga qo'shing yoki alohida fayl qilib import qiling
"""

import json
import csv
import io
from datetime import datetime

from django.contrib import admin, messages
from django.contrib.auth.admin import UserAdmin as BaseUserAdmin
from django.contrib.auth.hashers import make_password
from django.http import HttpResponse, HttpResponseRedirect
from django.shortcuts import render
from django.urls import path, reverse
from django.utils.html import format_html
from django import forms

# Modellarni import qiling (o'zingizning app nomingizga moslashtiring)
from .models import User, Region, District, Mahalla, Notification, Announcement


# ═══════════════════════════════════════════════════════════════════════════════
# JSON IMPORT FORMASI
# ═══════════════════════════════════════════════════════════════════════════════

class JSONImportForm(forms.Form):
    """JSON fayl yuklash formasi"""
    json_file = forms.FileField(
        label="JSON fayl",
        help_text="Yetakchilar ro'yxati bo'lgan JSON fayl yuklang"
    )

    generate_password_length = forms.IntegerField(
        label="Parol uzunligi",
        initial=8,
        min_value=6,
        max_value=20,
        help_text="Generatsiya qilinadigan parol uzunligi"
    )


# ═══════════════════════════════════════════════════════════════════════════════
# USER ADMIN - KENGAYTIRILGAN
# ═══════════════════════════════════════════════════════════════════════════════

@admin.register(User)
class UserAdmin(BaseUserAdmin):
    """Foydalanuvchi admin - JSON import bilan"""

    # ─────────────────────────────────────────────────────────────────────────
    # ASOSIY SOZLAMALAR
    # ─────────────────────────────────────────────────────────────────────────

    list_display = [
        'username', 'full_name_display', 'role', 'status',
        'mahalla_display', 'phone', 'plain_password_display', 'created_at'
    ]
    list_filter = ['role', 'status', 'region', 'district', 'created_at']
    search_fields = ['username', 'first_name', 'last_name', 'phone', 'mahalla__name']
    ordering = ['-created_at']

    list_per_page = 50

    fieldsets = (
        (None, {'fields': ('username', 'password')}),
        ('Shaxsiy ma\'lumotlar', {'fields': ('first_name', 'last_name', 'middle_name', 'phone', 'email')}),
        ('Joylashuv', {'fields': ('region', 'district', 'mahalla')}),
        ('Rol va holat', {'fields': ('role', 'status')}),
        ('Parol', {'fields': ('plain_password', 'must_change_password')}),
        ('Ruxsatlar', {'fields': ('is_active', 'is_staff', 'is_superuser', 'groups', 'user_permissions')}),
    )

    add_fieldsets = (
        (None, {
            'classes': ('wide',),
            'fields': ('username', 'password1', 'password2', 'first_name', 'last_name',
                       'phone', 'role', 'region', 'district', 'mahalla'),
        }),
    )

    readonly_fields = ['plain_password_display', 'created_at', 'updated_at']

    # ─────────────────────────────────────────────────────────────────────────
    # CUSTOM DISPLAY METHODS
    # ─────────────────────────────────────────────────────────────────────────

    @admin.display(description='F.I.O')
    def full_name_display(self, obj):
        return obj.full_name or '-'

    @admin.display(description='Mahalla')
    def mahalla_display(self, obj):
        if obj.mahalla:
            return obj.mahalla.name
        return '-'

    @admin.display(description='Parol (ochiq)')
    def plain_password_display(self, obj):
        if obj.plain_password:
            return format_html(
                '<code style="background:#f0f0f0;padding:2px 6px;border-radius:3px;">{}</code>',
                obj.plain_password
            )
        return '-'

    # ─────────────────────────────────────────────────────────────────────────
    # CUSTOM URLS
    # ─────────────────────────────────────────────────────────────────────────

    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path('import-json/', self.admin_site.admin_view(self.import_json_view), name='user_import_json'),
            path('export-credentials/', self.admin_site.admin_view(self.export_credentials_view),
                 name='user_export_credentials'),
            path('download-sample-json/', self.admin_site.admin_view(self.download_sample_json),
                 name='user_download_sample_json'),
        ]
        return custom_urls + urls

    # ─────────────────────────────────────────────────────────────────────────
    # CHANGE LIST - TUGMALAR QO'SHISH
    # ─────────────────────────────────────────────────────────────────────────

    def changelist_view(self, request, extra_context=None):
        extra_context = extra_context or {}
        extra_context['show_import_button'] = True
        extra_context['show_export_button'] = True
        return super().changelist_view(request, extra_context=extra_context)

    # ─────────────────────────────────────────────────────────────────────────
    # JSON IMPORT VIEW
    # ─────────────────────────────────────────────────────────────────────────

    def import_json_view(self, request):
        """JSON fayldan yetakchilarni import qilish"""

        if request.method == 'POST':
            form = JSONImportForm(request.POST, request.FILES)

            if form.is_valid():
                json_file = request.FILES['json_file']
                password_length = form.cleaned_data['generate_password_length']

                try:
                    # JSON o'qish
                    file_content = json_file.read().decode('utf-8')
                    data = json.loads(file_content)

                    if not isinstance(data, list):
                        messages.error(request, "JSON fayl ro'yxat formatida bo'lishi kerak!")
                        return HttpResponseRedirect(request.path)

                    # Import natijalarini saqlash
                    created_users = []
                    errors = []

                    for index, item in enumerate(data, start=1):
                        try:
                            result = self._create_user_from_json(item, password_length, request.user)
                            if result['success']:
                                created_users.append(result['user_data'])
                            else:
                                errors.append(f"Qator {index}: {result['error']}")
                        except Exception as e:
                            errors.append(f"Qator {index}: {str(e)}")

                    # Natijalarni session ga saqlash (export uchun)
                    if created_users:
                        request.session['imported_users'] = created_users
                        messages.success(
                            request,
                            f"✅ {len(created_users)} ta yetakchi muvaffaqiyatli qo'shildi!"
                        )

                    if errors:
                        error_text = "<br>".join(errors[:10])
                        if len(errors) > 10:
                            error_text += f"<br>... va yana {len(errors) - 10} ta xato"
                        messages.warning(request, format_html(f"⚠️ Xatolar:<br>{error_text}"))

                    # Export sahifasiga yo'naltirish
                    if created_users:
                        return HttpResponseRedirect(reverse('admin:user_export_credentials'))

                except json.JSONDecodeError as e:
                    messages.error(request, f"JSON format xatosi: {str(e)}")
                except Exception as e:
                    messages.error(request, f"Xatolik: {str(e)}")

                return HttpResponseRedirect(request.path)
        else:
            form = JSONImportForm()

        context = {
            'form': form,
            'title': 'JSON dan yetakchilarni import qilish',
            'opts': self.model._meta,
            'has_change_permission': True,
        }

        return render(request, 'admin/user_import_json.html', context)

    # ─────────────────────────────────────────────────────────────────────────
    # USER YARATISH (JSON dan)
    # ─────────────────────────────────────────────────────────────────────────

    def _create_user_from_json(self, item, password_length, created_by):
        """Bitta yetakchini JSON dan yaratish"""

        # Majburiy maydonlar
        first_name = item.get('first_name', '').strip()
        last_name = item.get('last_name', '').strip()

        if not first_name or not last_name:
            return {'success': False, 'error': 'first_name va last_name majburiy'}

        # Ixtiyoriy maydonlar
        middle_name = item.get('middle_name', '').strip()
        phone = item.get('phone', '').strip()
        email = item.get('email', '').strip()

        # Joylashuv
        region_name = item.get('region_name', '').strip() or item.get('region', '').strip()
        district_name = item.get('district_name', '').strip() or item.get('district', '').strip()
        mahalla_name = item.get('mahalla_name', '').strip() or item.get('mahalla', '').strip()

        # Region topish
        region = None
        if region_name:
            region = Region.objects.filter(name__iexact=region_name, is_active=True).first()
            if not region:
                # Qisman mos kelishni tekshirish
                region = Region.objects.filter(name__icontains=region_name, is_active=True).first()

        # District topish
        district = None
        if district_name:
            district_qs = District.objects.filter(name__iexact=district_name, is_active=True)
            if region:
                district_qs = district_qs.filter(region=region)
            district = district_qs.first()

            if not district:
                # Qisman mos kelishni tekshirish
                district_qs = District.objects.filter(name__icontains=district_name, is_active=True)
                if region:
                    district_qs = district_qs.filter(region=region)
                district = district_qs.first()

            # Districtdan regionni olish
            if district and not region:
                region = district.region

        # Mahalla topish
        mahalla = None
        if mahalla_name:
            mahalla_qs = Mahalla.objects.filter(name__iexact=mahalla_name, is_active=True)
            if district:
                mahalla_qs = mahalla_qs.filter(district=district)
            mahalla = mahalla_qs.first()

            if not mahalla:
                # Qisman mos kelishni tekshirish
                mahalla_qs = Mahalla.objects.filter(name__icontains=mahalla_name, is_active=True)
                if district:
                    mahalla_qs = mahalla_qs.filter(district=district)
                mahalla = mahalla_qs.first()

            # Mahalladan district va regionni olish
            if mahalla:
                if not district:
                    district = mahalla.district
                if not region:
                    region = mahalla.district.region

        # Telefon raqamni tekshirish (dublikat)
        if phone:
            if User.objects.filter(phone=phone).exists():
                return {'success': False, 'error': f'Bu telefon raqam allaqachon mavjud: {phone}'}

        # Username generatsiya
        username = User.generate_username(first_name, last_name)

        # Parol generatsiya
        plain_password = User.generate_password(password_length)

        # User yaratish
        user = User(
            username=username,
            first_name=first_name,
            last_name=last_name,
            middle_name=middle_name,
            phone=phone if phone else None,
            email=email if email else None,
            region=region,
            district=district,
            mahalla=mahalla,
            role=User.Role.LEADER,
            status=User.Status.ACTIVE,
            plain_password=plain_password,
            must_change_password=True,
            created_by=created_by,
        )

        # Parolni hash qilish
        user.password = make_password(plain_password)
        user.save()

        # Natija
        return {
            'success': True,
            'user_data': {
                'id': str(user.id),
                'username': username,
                'password': plain_password,
                'full_name': user.full_name,
                'first_name': first_name,
                'last_name': last_name,
                'middle_name': middle_name,
                'phone': phone or '-',
                'email': email or '-',
                'region': region.name if region else '-',
                'district': district.name if district else '-',
                'mahalla': mahalla.name if mahalla else '-',
            }
        }

    # ─────────────────────────────────────────────────────────────────────────
    # EXPORT CREDENTIALS VIEW
    # ─────────────────────────────────────────────────────────────────────────

    def export_credentials_view(self, request):
        """Import qilingan userlarning login/parollarini ko'rsatish va yuklab olish"""

        imported_users = request.session.get('imported_users', [])

        # CSV yuklab olish
        if request.GET.get('format') == 'csv':
            return self._export_csv(imported_users)

        # Excel yuklab olish
        if request.GET.get('format') == 'excel':
            return self._export_excel(imported_users)

        context = {
            'title': 'Import qilingan yetakchilar',
            'users': imported_users,
            'total': len(imported_users),
            'opts': self.model._meta,
            'has_change_permission': True,
        }

        return render(request, 'admin/user_export_credentials.html', context)

    def _export_csv(self, users):
        """CSV formatda eksport"""
        response = HttpResponse(content_type='text/csv; charset=utf-8')
        response[
            'Content-Disposition'] = f'attachment; filename="yetakchilar_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv"'

        # UTF-8 BOM qo'shish (Excel uchun)
        response.write('\ufeff')

        writer = csv.writer(response)
        writer.writerow(['#', 'F.I.O', 'Login', 'Parol', 'Telefon', 'Viloyat', 'Tuman', 'Mahalla'])

        for i, user in enumerate(users, start=1):
            writer.writerow([
                i,
                user['full_name'],
                user['username'],
                user['password'],
                user['phone'],
                user['region'],
                user['district'],
                user['mahalla'],
            ])

        return response

    def _export_excel(self, users):
        """Excel formatda eksport (openpyxl kerak)"""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        except ImportError:
            # openpyxl yo'q bo'lsa CSV qaytarish
            return self._export_csv(users)

        wb = Workbook()
        ws = wb.active
        ws.title = "Yetakchilar"

        # Sarlavha stili
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # Chegara
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Sarlavhalar
        headers = ['#', 'F.I.O', 'Login', 'Parol', 'Telefon', 'Viloyat', 'Tuman', 'Mahalla']

        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        # Ma'lumotlar
        for row, user in enumerate(users, start=2):
            data = [
                row - 1,
                user['full_name'],
                user['username'],
                user['password'],
                user['phone'],
                user['region'],
                user['district'],
                user['mahalla'],
            ]

            for col, value in enumerate(data, start=1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.border = thin_border
                cell.alignment = Alignment(vertical="center")

        # Ustun kengliklari
        column_widths = [5, 30, 20, 15, 18, 20, 20, 25]
        for i, width in enumerate(column_widths, start=1):
            ws.column_dimensions[chr(64 + i)].width = width

        # Response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response[
            'Content-Disposition'] = f'attachment; filename="yetakchilar_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'

        wb.save(response)
        return response

    # ─────────────────────────────────────────────────────────────────────────
    # NAMUNA JSON YUKLAB OLISH
    # ─────────────────────────────────────────────────────────────────────────

    def download_sample_json(self, request):
        """Namuna JSON fayl yuklab olish"""

        sample_data = [
            {
                "first_name": "Shermat",
                "last_name": "Shermatov",
                "middle_name": "Shermatovich",
                "phone": "+998901234567",
                "email": "shermat@example.com",
                "region_name": "Toshkent shahri",
                "district_name": "Chilonzor tumani",
                "mahalla_name": "Tinchlik MFY"
            },
            {
                "first_name": "Olim",
                "last_name": "Olimov",
                "middle_name": "Olimovich",
                "phone": "+998901234568",
                "region_name": "Toshkent shahri",
                "district_name": "Yakkasaroy tumani",
                "mahalla_name": "Bobur MFY"
            },
            {
                "first_name": "Karim",
                "last_name": "Karimov",
                "middle_name": "Karimovich",
                "phone": "+998901234569",
                "region_name": "Samarqand viloyati",
                "district_name": "Samarqand shahri",
                "mahalla_name": "Registon MFY"
            }
        ]

        response = HttpResponse(
            json.dumps(sample_data, ensure_ascii=False, indent=2),
            content_type='application/json; charset=utf-8'
        )
        response['Content-Disposition'] = 'attachment; filename="namuna_yetakchilar.json"'

        change_list_template = 'admin/user_change_list.html'  # ← Qo'shin

        return response


# ═══════════════════════════════════════════════════════════════════════════════
# BOSHQA MODELLAR ADMIN
# ═══════════════════════════════════════════════════════════════════════════════

@admin.register(Region)
class RegionAdmin(admin.ModelAdmin):
    list_display = ['name', 'code', 'districts_count', 'leaders_count', 'is_active', 'order']
    list_filter = ['is_active']
    search_fields = ['name', 'code']
    ordering = ['order', 'name']
    list_editable = ['order', 'is_active']


@admin.register(District)
class DistrictAdmin(admin.ModelAdmin):
    list_display = ['name', 'region', 'code', 'mahallas_count', 'leaders_count', 'is_active', 'order']
    list_filter = ['region', 'is_active']
    search_fields = ['name', 'code', 'region__name']
    ordering = ['region', 'order', 'name']
    list_editable = ['order', 'is_active']
    autocomplete_fields = ['region']


@admin.register(Mahalla)
class MahallaAdmin(admin.ModelAdmin):
    list_display = ['name', 'district', 'region_display', 'population', 'has_leader_display', 'is_active']
    list_filter = ['district__region', 'district', 'is_active']
    search_fields = ['name', 'district__name', 'district__region__name']
    ordering = ['district', 'name']
    autocomplete_fields = ['district']

    @admin.display(description='Viloyat')
    def region_display(self, obj):
        return obj.district.region.name

    @admin.display(description='Yetakchi', boolean=True)
    def has_leader_display(self, obj):
        return obj.has_leader


@admin.register(Notification)
class NotificationAdmin(admin.ModelAdmin):
    list_display = ['user', 'type', 'title', 'is_read', 'created_at']
    list_filter = ['type', 'is_read', 'priority', 'created_at']
    search_fields = ['title', 'message', 'user__username']
    ordering = ['-created_at']
    readonly_fields = ['created_at', 'read_at']


@admin.register(Announcement)
class AnnouncementAdmin(admin.ModelAdmin):
    list_display = ['title', 'priority', 'status', 'target_display', 'views_count', 'created_at']
    list_filter = ['status', 'priority', 'target_all', 'target_region']
    search_fields = ['title', 'content']
    ordering = ['-created_at']

    @admin.display(description='Maqsad')
    def target_display(self, obj):
        if obj.target_all:
            return "Hammaga"
        parts = []
        if obj.target_region:
            parts.append(obj.target_region.name)
        if obj.target_district:
            parts.append(obj.target_district.name)
        return ", ".join(parts) if parts else "-"